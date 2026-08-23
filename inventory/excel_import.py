"""
Parsing/import logic for the "upload the Excel file of the inventory"
feature from the original brief. Kept out of views.py because it's a
decent chunk of row-by-row logic that has nothing to do with request
handling.

Expected columns (case-insensitive, any order — matched by header name,
not position): SKU, Name, Description, Track Quantity, Unit, Quantity,
Location, Location Description, Reorder Level, Allow Take, Allow Borrow.
Only SKU and Name are required; everything else falls back to the same
defaults the "Add item" form uses if the column is missing or the cell
is blank.
"""

import openpyxl

from .models import Item, StockTransaction

REQUIRED_HEADERS = {'sku', 'name'}

# Maps a normalized header (lowercased, stripped) to the field name we use
# internally. Accepts a couple of friendly variants per column.
HEADER_ALIASES = {
    'sku': 'sku',
    'name': 'name',
    'description': 'description',
    'track quantity': 'track_quantity',
    'track_quantity': 'track_quantity',
    'unit': 'unit',
    'quantity': 'quantity',
    'qty': 'quantity',
    'location': 'location',
    'location description': 'location_description',
    'location_description': 'location_description',
    'reorder level': 'reorder_level',
    'reorder_level': 'reorder_level',
    'allow take': 'allow_take',
    'allow_take': 'allow_take',
    'allow borrow': 'allow_borrow',
    'allow_borrow': 'allow_borrow',
}

TRUE_STRINGS = {'yes', 'y', 'true', '1'}
FALSE_STRINGS = {'no', 'n', 'false', '0'}


def _parse_bool(value, default):
    if value is None or str(value).strip() == '':
        return default
    text = str(value).strip().lower()
    if text in TRUE_STRINGS:
        return True
    if text in FALSE_STRINGS:
        return False
    return default


def _parse_int(value, default):
    if value is None or str(value).strip() == '':
        return default
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return default


class ExcelImportError(Exception):
    """Raised for a problem with the file itself (not a single bad row)."""


def import_items_from_excel(file_obj, organization, user):
    """
    Reads an uploaded .xlsx file and creates/updates Items for the given
    organization. Returns a dict: {'created': int, 'updated': int,
    'skipped': [(row_number, reason), ...]}.

    A SKU that doesn't exist yet becomes a new item, with the row's
    quantity as its starting stock. A SKU that already exists is treated
    like a restock: its details (name, description, unit, location,
    reorder level, take/borrow flags) are refreshed from the row, and the
    row's quantity is ADDED to whatever stock it already had — not
    replaced — since a re-uploaded spreadsheet usually represents a new
    delivery, not a correction. Every row that touches an item leaves a
    "Bulk Excel import" entry in that item's history, same as any other
    stock change.

    Location is plain text on the item itself, not a shared object — a
    row's Location/Location Description columns only ever set that one
    row's item, with no effect on any other item that happens to share
    the same location name.
    """
    try:
        workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError(
            f"Couldn't read that file as an Excel workbook ({exc}). Make sure it's a real .xlsx file."
        )

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise ExcelImportError('That file looks empty — no header row found.')

    column_map = {}  # column index -> internal field name
    for index, raw_header in enumerate(header_row):
        if raw_header is None:
            continue
        normalized = str(raw_header).strip().lower()
        field = HEADER_ALIASES.get(normalized)
        if field:
            column_map[index] = field

    found_fields = set(column_map.values())
    missing_required = REQUIRED_HEADERS - found_fields
    if missing_required:
        raise ExcelImportError(
            f"Missing required column(s): {', '.join(sorted(missing_required))}. "
            "Download the template below to see the exact headers expected."
        )

    created = 0
    updated = 0
    skipped = []

    for row_number, row in enumerate(rows, start=2):  # row 1 was the header
        if row is None or all(cell in (None, '') for cell in row):
            continue  # silently skip fully blank rows

        values = {}
        for index, field in column_map.items():
            values[field] = row[index] if index < len(row) else None

        sku = str(values.get('sku') or '').strip()
        name = str(values.get('name') or '').strip()
        if not sku or not name:
            skipped.append((row_number, 'missing SKU or Name'))
            continue

        location_name = str(values.get('location') or '').strip()
        location_description = str(values.get('location_description') or '').strip()

        item = Item.objects.filter(organization=organization, sku=sku).first()

        track_quantity = _parse_bool(
            values.get('track_quantity'),
            default=item.track_quantity if item is not None else True,
        )
        # Quantity/reorder level/take/borrow don't mean anything for an item
        # that isn't quantity-tracked (bulk material, location-only) — same
        # rule the "Add item" form enforces — so a row that turns tracking
        # off ignores whatever those columns say.
        row_quantity = 0 if not track_quantity else _parse_int(values.get('quantity'), default=0)

        if item is None:
            item = Item.objects.create(
                organization=organization,
                sku=sku,
                name=name,
                description=str(values.get('description') or ''),
                track_quantity=track_quantity,
                unit=str(values.get('unit') or 'pcs').strip() or 'pcs',
                quantity=row_quantity,
                location_name=location_name,
                location_description=location_description,
                reorder_level=(
                    0 if not track_quantity else _parse_int(values.get('reorder_level'), default=0)
                ),
                allow_take=track_quantity and _parse_bool(values.get('allow_take'), default=True),
                allow_borrow=track_quantity and _parse_bool(values.get('allow_borrow'), default=False),
            )
            StockTransaction.objects.create(
                item=item,
                transaction_type=StockTransaction.EXCEL_IMPORT,
                quantity_change=row_quantity,
                new_location_name=location_name,
                performed_by=user,
                note='Created via Excel import',
            )
            created += 1
        else:
            item.name = name
            item.description = str(values.get('description') or item.description)
            item.track_quantity = track_quantity
            if values.get('unit'):
                item.unit = str(values['unit']).strip()
            # Blank Location/Location Description cells leave the item's
            # existing values alone rather than clearing them — a row that
            # only means to restock quantity shouldn't accidentally wipe
            # out a location that was set some other way.
            previous_location_name = item.location_name
            if location_name:
                item.location_name = location_name
            if location_description:
                item.location_description = location_description
            if not track_quantity:
                item.reorder_level = 0
                item.allow_take = False
                item.allow_borrow = False
            else:
                item.reorder_level = _parse_int(values.get('reorder_level'), default=item.reorder_level)
                item.allow_take = _parse_bool(values.get('allow_take'), default=item.allow_take)
                item.allow_borrow = _parse_bool(values.get('allow_borrow'), default=item.allow_borrow)
            item.quantity += row_quantity
            item.save()
            StockTransaction.objects.create(
                item=item,
                transaction_type=StockTransaction.EXCEL_IMPORT,
                quantity_change=row_quantity,
                previous_location_name=previous_location_name,
                new_location_name=item.location_name,
                performed_by=user,
                note='Updated via Excel import',
            )
            updated += 1

    return {'created': created, 'updated': updated, 'skipped': skipped}


def build_template_workbook():
    """Builds an in-memory .xlsx with the expected headers and one example row."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Inventory'
    sheet.append([
        'SKU', 'Name', 'Description', 'Track Quantity', 'Unit', 'Quantity', 'Location',
        'Location Description', 'Reorder Level', 'Allow Take', 'Allow Borrow',
    ])
    sheet.append([
        'DRILL-001', 'Cordless drill', '18V, comes with 2 batteries', 'Yes', 'pcs', 10,
        'Main warehouse', 'Aisle 3, shelf B', 2, 'No', 'Yes',
    ])
    sheet.append([
        'SAND-BULK', 'Sand (bulk pile)', 'Not counted individually', 'No', '', '',
        'Yard 2', 'Behind the main building', '', '', '',
    ])
    return workbook
